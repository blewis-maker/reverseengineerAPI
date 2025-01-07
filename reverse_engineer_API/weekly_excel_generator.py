import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

# Configure logger
logger = logging.getLogger(__name__)

class WeeklyReportGenerator:
    def __init__(self, metrics, report_date=None):
        self.metrics = metrics
        self.current_row = 1  # Initialize current_row
        # Ensure report_date is a datetime object
        if isinstance(report_date, str):
            try:
                self.report_date = datetime.strptime(report_date, '%Y-%m-%d')
            except ValueError:
                self.report_date = datetime.now()
        elif isinstance(report_date, datetime):
            self.report_date = report_date
        else:
            self.report_date = datetime.now()
            
        # Adjust report_date to Sunday 8:00 AM if not already
        days_until_sunday = (6 - self.report_date.weekday()) % 7
        self.report_date = self.report_date + timedelta(days=days_until_sunday)
        self.report_date = self.report_date.replace(hour=8, minute=0, second=0, microsecond=0)
        
        logger.info(f"Initializing Weekly Report Generator for date: {self.report_date}")
        self.wb = Workbook()
        
    def generate_report(self, output_path):
        """Generate the weekly report Excel file."""
        try:
            logger.info(f"Starting weekly report generation at {output_path}")
            logger.info(f"Report period: {(self.report_date - timedelta(days=7)).strftime('%Y-%m-%d')} to {self.report_date.strftime('%Y-%m-%d')}")
            
            # Create sheets
            weekly_metrics = self.wb.active
            weekly_metrics.title = "Weekly Status"
            burndown = self.wb.create_sheet("Burndown")
            schedule = self.wb.create_sheet("Schedule")
            
            logger.info("Created worksheet structure")
            
            # Get weekly status once
            weekly_status = self.metrics.get_weekly_status()
            
            # Generate each section
            logger.info("Generating report sections...")
            self._generate_header(weekly_metrics)
            
            # Generate utility progress
            logger.info("Generating Utility Progress section")
            self._generate_utility_progress(weekly_metrics, weekly_status)
            
            # Generate OSP productivity
            logger.info("Generating OSP Productivity section")
            self._generate_osp_productivity(weekly_metrics, weekly_status)
            
            # Generate status tracking
            logger.info("Generating Status Tracking section")
            self._generate_status_tracking(weekly_metrics, weekly_status)
            
            # Generate burndown metrics
            logger.info("Generating Burndown Metrics section")
            self._generate_burndown_metrics(burndown, weekly_status)
            
            # Generate schedule metrics
            logger.info("Generating Schedule Metrics section")
            self._generate_schedule_metrics(schedule, weekly_status)
            
            # Save the workbook
            self.wb.save(output_path)
            logger.info(f"Report saved to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error generating report: {str(e)}")
            logger.error("Failed to generate report")
            return False
            
    def _generate_header(self, ws):
        """Generate the report header."""
        # Title
        ws['A1'] = "OSP Production Master - Weekly Status Report"
        ws['A1'].font = Font(bold=True)
        
        # Date range in merged cells
        start_date = self.report_date - timedelta(days=7)
        date_range = f"Date Range: {start_date.strftime('%Y-%m-%d')} to {self.report_date.strftime('%Y-%m-%d')}"
        ws['B1'] = date_range
        ws.merge_cells('B1:D1')
        ws['B1'].font = Font(bold=True)
        ws['B1'].alignment = Alignment(horizontal='center')
        
        # Add spacing for the next section
        self.current_row = 3
        
    def _generate_status_tracking(self, worksheet, weekly_status):
        """Generate status tracking section of the report."""
        logging.info("Generating status tracking section")
        
        # Section header
        worksheet.cell(row=self.current_row, column=1).value = "Status Tracking"
        worksheet.cell(row=self.current_row, column=1).font = Font(bold=True)
        self.current_row += 1
        
        # Set up headers
        headers = ['Status', 'Total Jobs', 'Total Poles', 'Change from Last Week']
        for col, header in enumerate(headers):
            cell = worksheet.cell(row=self.current_row, column=col + 1)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
        self.current_row += 1
        
        # Add status changes
        status_changes = weekly_status.get('status_changes', {})
        for status, data in status_changes.items():
            row = [
                status,
                data.get('job_count', 0),
                data.get('pole_count', 0),
                f"+{data.get('change_from_last_week', 0)}" if data.get('change_from_last_week', 0) > 0 else str(data.get('change_from_last_week', 0))
            ]
            for col, value in enumerate(row):
                cell = worksheet.cell(row=self.current_row, column=col + 1)
                cell.value = value
            self.current_row += 1
                
        logging.info(f"Added status tracking data")
        
        # Add spacing
        self.current_row += 2

    def _generate_burndown_metrics(self, worksheet, weekly_status):
        """Generate enhanced burndown metrics section with multiple views."""
        logging.info("Generating burndown metrics section")
        
        # Add master burndown header
        self._add_section_header(worksheet, "Master Burndown", self.current_row)
        self.current_row += 1
        
        # Set up headers for master burndown
        headers = ['Total Poles', 'Field Complete', 'Back Office Complete', 'Overall Progress', 'Current Run Rate', 'Est. Completion']
        self._add_headers(worksheet, headers, self.current_row)
        self.current_row += 1
        
        # Get master burndown metrics
        master_metrics = weekly_status['burndown']['master']
        total_poles = master_metrics.get('total_poles', 0)
        field_complete = master_metrics.get('field_complete', 0)
        back_office_complete = master_metrics.get('back_office_complete', 0)
        overall_progress = (back_office_complete / total_poles * 100) if total_poles > 0 else 0
        run_rate = master_metrics.get('run_rate', 0)
        est_completion = master_metrics.get('estimated_completion_date')
        
        # Add master burndown row
        row = [
            total_poles,
            field_complete,
            back_office_complete,
            f"{overall_progress:.1f}%",
            f"{run_rate:.1f} poles/week",
            est_completion.strftime('%Y-%m-%d') if est_completion else 'TBD'
        ]
        self._add_row(worksheet, row, self.current_row)
        self.current_row += 2
        
        # Add utility burndown header
        self._add_section_header(worksheet, "Utility Burndown", self.current_row)
        self.current_row += 1
        
        # Set up headers for utility burndown
        headers = ['Utility', 'Total Poles', 'Field Complete', 'Back Office Complete', 'Progress', 'Run Rate', 'Est. Completion']
        self._add_headers(worksheet, headers, self.current_row)
        self.current_row += 1
        
        # Add utility burndown data
        utility_metrics = weekly_status['burndown']['by_utility']
        utility_start_row = self.current_row
        for utility, metrics in utility_metrics.items():
            if utility == 'Unknown':
                continue
                
            progress = (metrics['back_office_complete'] / metrics['total_poles'] * 100) if metrics['total_poles'] > 0 else 0
            row = [
                utility,
                metrics['total_poles'],
                metrics['field_complete'],
                metrics['back_office_complete'],
                f"{progress:.1f}%",
                f"{metrics['run_rate']:.1f} poles/week",
                metrics['estimated_completion'].strftime('%Y-%m-%d') if metrics.get('estimated_completion') else 'TBD'
            ]
            self._add_row(worksheet, row, self.current_row)
            self.current_row += 1
        utility_end_row = self.current_row - 1
        
        # Add project burndown header
        self.current_row += 2
        self._add_section_header(worksheet, "Project Burndown", self.current_row)
        self.current_row += 1
        
        # Set up headers for project burndown
        headers = ['Project', 'Total Poles', 'Field Complete', 'Back Office Complete', 'Progress', 'Run Rate', 'Target Date', 'Est. Completion', 'Status']
        self._add_headers(worksheet, headers, self.current_row)
        self.current_row += 1
        
        # Add project burndown data
        project_metrics = weekly_status['burndown']['by_project']
        project_start_row = self.current_row
        for project, metrics in project_metrics.items():
            if project == 'Unknown':
                continue
                
            progress = (metrics['back_office_complete'] / metrics['total_poles'] * 100) if metrics['total_poles'] > 0 else 0
            target_date = metrics.get('target_date')
            est_completion = metrics.get('estimated_completion')
            
            # Calculate status
            status = 'Not Started'
            status_color = 'FFFFFF'  # White
            if est_completion:
                if target_date:
                    if est_completion > target_date:
                        status = 'Behind'
                        status_color = 'FF6B6B'  # Light red
                    elif (est_completion - target_date).days <= -14:
                        status = 'On Track'
                        status_color = '90EE90'  # Light green
                    else:
                        status = 'At Risk'
                        status_color = 'FFD700'  # Gold
            
            row = [
                project,
                metrics['total_poles'],
                metrics['field_complete'],
                metrics['back_office_complete'],
                f"{progress:.1f}%",
                f"{metrics['run_rate']:.1f} poles/week",
                target_date.strftime('%Y-%m-%d') if target_date else 'TBD',
                est_completion.strftime('%Y-%m-%d') if est_completion else 'TBD',
                status
            ]
            self._add_row(worksheet, row, self.current_row, status_color=status_color)
            self.current_row += 1
        project_end_row = self.current_row - 1
        
        # Add spacing
        self.current_row += 2
        
        # Create burndown charts
        self._create_burndown_chart(worksheet, utility_start_row, utility_end_row, "Utility Burndown", "I2")
        self._create_burndown_chart(worksheet, project_start_row, project_end_row, "Project Burndown", "I20")
        
        # Add trend analysis
        self.current_row += 2
        self._add_section_header(worksheet, "7-Day Trend Analysis", self.current_row)
        self.current_row += 1
        
        headers = ['Entity', 'Last 7 Days Progress', 'Trend', 'Projected Completion']
        self._add_headers(worksheet, headers, self.current_row)
        self.current_row += 1
        
        # Add master trend
        master_trend = self._calculate_trend(master_metrics.get('history', []))
        self._add_row(worksheet, [
            'Overall',
            f"{master_trend['progress']:.1f}%",
            master_trend['indicator'],
            master_trend['projection'].strftime('%Y-%m-%d') if master_trend.get('projection') else 'TBD'
        ], self.current_row)
        self.current_row += 1
        
        # Add utility trends
        for utility, metrics in utility_metrics.items():
            if utility == 'Unknown':
                continue
            trend = self._calculate_trend(metrics.get('history', []))
            self._add_row(worksheet, [
                utility,
                f"{trend['progress']:.1f}%",
                trend['indicator'],
                trend['projection'].strftime('%Y-%m-%d') if trend.get('projection') else 'TBD'
            ], self.current_row)
            self.current_row += 1
        
        # Add project trends
        for project, metrics in project_metrics.items():
            if project == 'Unknown':
                continue
            trend = self._calculate_trend(metrics.get('history', []))
            self._add_row(worksheet, [
                project,
                f"{trend['progress']:.1f}%",
                trend['indicator'],
                trend['projection'].strftime('%Y-%m-%d') if trend.get('projection') else 'TBD'
            ], self.current_row)
            self.current_row += 1

    def _calculate_trend(self, history):
        """Calculate trend from historical data."""
        if not history or len(history) < 2:
            return {'progress': 0, 'indicator': '→', 'projection': None}
        
        # Sort history by timestamp
        history = sorted(history, key=lambda x: x['timestamp'])
        
        # Get last 7 days of data
        week_ago = history[-1]['timestamp'] - timedelta(days=7)
        week_data = [h for h in history if h['timestamp'] >= week_ago]
        
        if not week_data:
            return {'progress': 0, 'indicator': '→', 'projection': None}
        
        # Calculate progress
        start_complete = week_data[0]['completed_poles']
        end_complete = week_data[-1]['completed_poles']
        progress = ((end_complete - start_complete) / start_complete * 100) if start_complete > 0 else 0
        
        # Determine trend
        if progress > 5:
            indicator = '↑'  # Increasing
        elif progress < -5:
            indicator = '↓'  # Decreasing
        else:
            indicator = '→'  # Stable
        
        # Calculate projection
        if len(week_data) >= 2:
            daily_rate = (end_complete - start_complete) / 7
            if daily_rate > 0:
                remaining = history[-1]['total_poles'] - end_complete
                days_to_complete = remaining / daily_rate
                projection = history[-1]['timestamp'] + timedelta(days=days_to_complete)
            else:
                projection = None
        else:
            projection = None
        
        return {
            'progress': progress,
            'indicator': indicator,
            'projection': projection
        }

    def _generate_schedule_metrics(self, worksheet, weekly_status):
        """Generate enhanced schedule metrics section."""
        logging.info("Generating schedule metrics section")
        
        # Add schedule header
        self._add_section_header(worksheet, "Project Schedule", self.current_row)
        self.current_row += 1
        
        # Set up headers
        headers = [
            'Project', 'Total Poles', 'Progress', 'Field Resources', 'Back Office Resources',
            'Current Run Rate', 'Required Run Rate', 'Target Date', 'Projected End Date', 'Status'
        ]
        self._add_headers(worksheet, headers, self.current_row)
        self.current_row += 1
        
        # Add project schedule data
        schedule = weekly_status['schedule']
        for project in schedule['projects']:
            progress = (project['completed_poles'] / project['total_poles'] * 100) if project['total_poles'] > 0 else 0
            
            # Calculate required run rate
            if project.get('target_date'):
                remaining_poles = project['total_poles'] - project['completed_poles']
                remaining_weeks = (project['target_date'] - datetime.now()).days / 7
                required_rate = remaining_poles / remaining_weeks if remaining_weeks > 0 else float('inf')
            else:
                required_rate = 0
            
            # Determine status
            status = project.get('status', 'Not Started')
            status_color = {
                'On Track': '90EE90',  # Light green
                'At Risk': 'FFD700',   # Gold
                'Behind': 'FF6B6B',    # Light red
                'Not Started': 'FFFFFF' # White
            }.get(status, 'FFFFFF')
            
            row = [
                project['project_id'],
                project['total_poles'],
                f"{progress:.1f}%",
                len(project['resources']['field']),
                len(project['resources']['back_office']),
                f"{project['run_rate']:.1f} poles/week",
                f"{required_rate:.1f} poles/week",
                project['target_date'].strftime('%Y-%m-%d') if project.get('target_date') else 'TBD',
                project['estimated_completion'].strftime('%Y-%m-%d') if project.get('estimated_completion') else 'TBD',
                status
            ]
            self._add_row(worksheet, row, self.current_row, status_color=status_color)
            self.current_row += 1
        
        # Create Gantt chart
        self._create_gantt_chart(worksheet, schedule['projects'])

    def _create_gantt_chart(self, worksheet, projects):
        """Create a Gantt chart for project schedule visualization."""
        # Start Gantt chart 2 rows below current position
        self.current_row += 2
        self._add_section_header(worksheet, "Project Timeline (Gantt Chart)", self.current_row)
        self.current_row += 1
        
        # Calculate date range
        start_date = min((p['start_date'] for p in projects if p.get('start_date')), default=datetime.now())
        end_date = max((p['estimated_completion'] for p in projects if p.get('estimated_completion')), 
                      default=datetime.now() + timedelta(days=90))
        
        # Create timeline header
        timeline_start = start_date.replace(day=1)
        timeline_end = (end_date.replace(day=1) + timedelta(days=32)).replace(day=1)
        current_date = timeline_start
        col = 2
        
        while current_date < timeline_end:
            cell = worksheet.cell(row=self.current_row, column=col)
            cell.value = current_date.strftime('%Y-%m')
            cell.font = Font(bold=True)
            current_date = (current_date + timedelta(days=32)).replace(day=1)
            col += 1
        
        self.current_row += 1
        
        # Add project timelines
        for project in projects:
            # Project name
            worksheet.cell(row=self.current_row, column=1).value = project['project_id']
            
            if project.get('start_date') and project.get('estimated_completion'):
                start_col = 2 + (project['start_date'].year - timeline_start.year) * 12 + \
                           (project['start_date'].month - timeline_start.month)
                duration_months = (project['estimated_completion'].year - project['start_date'].year) * 12 + \
                                (project['estimated_completion'].month - project['start_date'].month) + 1
                
                # Draw progress bar
                for i in range(duration_months):
                    cell = worksheet.cell(row=self.current_row, column=start_col + i)
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    
                    # Add completion percentage if applicable
                    if project['total_poles'] > 0:
                        progress = project['completed_poles'] / project['total_poles']
                        if i < int(duration_months * progress):
                            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            
            self.current_row += 1

    def _add_section_header(self, worksheet, title, row):
        """Add a section header with consistent formatting."""
        cell = worksheet.cell(row=row, column=1)
        cell.value = title
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

    def _add_headers(self, worksheet, headers, row):
        """Add headers with consistent formatting."""
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

    def _add_row(self, worksheet, values, row, status_color=None):
        """Add a row with consistent formatting."""
        for col, value in enumerate(values, 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal='center')
            if status_color and col == len(values):  # Color the status column if color provided
                cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')

    def _generate_utility_progress(self, worksheet, weekly_status):
        """Generate utility progress section."""
        # Section header
        cell = worksheet.cell(row=4, column=1)
        cell.value = "Utility Progress"
        cell.font = Font(bold=True)
        
        # Write headers
        headers = ['Utility', 'Total Poles', 'Field Completed', 'Back Office Completed', 'Remaining Poles', 'Run Rate', 'Est. Completion']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
        row = 6
        for utility, data in weekly_status['utility_metrics'].items():
            worksheet.cell(row=row, column=1).value = utility
            worksheet.cell(row=row, column=2).value = data['total_poles']
            worksheet.cell(row=row, column=3).value = data['field_completed']
            worksheet.cell(row=row, column=4).value = data['back_office_completed']
            remaining = data['total_poles'] - data['field_completed'] - data['back_office_completed']
            worksheet.cell(row=row, column=5).value = remaining
            worksheet.cell(row=row, column=6).value = data['run_rate']
            worksheet.cell(row=row, column=7).value = data['estimated_completion']
            row += 1
            
        self.current_row = row + 2  # Update current row with spacing

    def _generate_osp_productivity(self, worksheet, weekly_status):
        """Generate OSP productivity section."""
        self.current_row = self._write_field_productivity(worksheet, weekly_status['user_production']['field'])
        self.current_row = self._write_back_office_productivity(worksheet, weekly_status['user_production'])

    def _write_field_productivity(self, worksheet, field_users):
        """Write field productivity section."""
        # Section header
        cell = worksheet.cell(row=self.current_row, column=1)
        cell.value = "Field Productivity"
        cell.font = Font(bold=True)
        self.current_row += 1
        
        # Write headers
        headers = ['User', 'Completed Poles', 'Utilities', 'Jobs Completed']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=self.current_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
        self.current_row += 1
        
        for user_data in field_users:
            worksheet.cell(row=self.current_row, column=1).value = user_data['user']
            worksheet.cell(row=self.current_row, column=2).value = user_data['completed_poles']
            worksheet.cell(row=self.current_row, column=3).value = ', '.join(user_data['utilities'])
            jobs_text = ', '.join([f"{job['job_id']} ({job['pole_count']} poles)" for job in user_data['jobs']])
            worksheet.cell(row=self.current_row, column=4).value = jobs_text
            self.current_row += 1
            
        return self.current_row + 1

    def _write_back_office_productivity(self, worksheet, user_production):
        """Write back office productivity section."""
        # Section header
        cell = worksheet.cell(row=self.current_row, column=1)
        cell.value = "Back Office Productivity"
        cell.font = Font(bold=True)
        self.current_row += 1
        
        # Write headers
        headers = ['Category', 'User', 'Completed Poles', 'Utilities', 'Jobs Completed']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=self.current_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
        self.current_row += 1
        
        categories = ['annotation', 'sent_to_pe', 'delivery', 'emr', 'approved']
        for category in categories:
            for user_data in user_production[category]:
                worksheet.cell(row=self.current_row, column=1).value = category.replace('_', ' ').title()
                worksheet.cell(row=self.current_row, column=2).value = user_data['user']
                worksheet.cell(row=self.current_row, column=3).value = user_data['completed_poles']
                worksheet.cell(row=self.current_row, column=4).value = ', '.join(user_data['utilities'])
                jobs_text = ', '.join([f"{job['job_id']} ({job['pole_count']} poles)" for job in user_data['jobs']])
                worksheet.cell(row=self.current_row, column=5).value = jobs_text
                self.current_row += 1
                
        return self.current_row + 1

    def _format_worksheet(self, ws):
        """Apply consistent formatting to the worksheet."""
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Center align all cells
        for row in ws.rows:
            for cell in row:
                cell.alignment = Alignment(horizontal='center') 

    def _create_burndown_chart(self, worksheet, start_row, end_row, title, position):
        """Create a burndown chart for visualization."""
        chart = BarChart()
        chart.title = title
        chart.style = 10
        chart.x_axis.title = "Entity"
        chart.y_axis.title = "Poles"
        
        # Add data series for total and completed poles
        data = Reference(worksheet, min_col=2, max_col=4,  # Include field and back office completion
                        min_row=start_row - 1, max_row=end_row)
        cats = Reference(worksheet, min_col=1, max_col=1,
                        min_row=start_row, max_row=end_row)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Set colors for the series
        chart.series[0].graphicalProperties.solidFill = "4472C4"  # Total poles - blue
        chart.series[1].graphicalProperties.solidFill = "ED7D31"  # Field complete - orange
        chart.series[2].graphicalProperties.solidFill = "90EE90"  # Back office complete - green
        
        worksheet.add_chart(chart, position) 