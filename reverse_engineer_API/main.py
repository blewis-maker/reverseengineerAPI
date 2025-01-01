import json
import http.client
import geopandas as gpd
from shapely.geometry import Point, LineString
from shapely.geometry import mapping
import os
import time
import socket
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import logging
from dotenv import load_dotenv
import msal
from O365 import Account, FileSystemTokenBackend
import requests
import base64
import warnings
import zipfile
from io import BytesIO
warnings.filterwarnings('ignore')

# Add at start of script
logging.basicConfig(
    filename='katapult_automation.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Load environment variables from .env file
load_dotenv()

# Toggle to enable/disable testing a specific job
TEST_ONLY_SPECIFIC_JOB = True

# ID of the specific job to test
TEST_JOB_ID = "-O-nlOLQbPIYhHwJCPDN"

# Add at top of script
CONFIG = {
    'API_KEY': 'rt2JR8Rds03Ry03hQTpD9j0N01gWEULJnuY3l1_GeXA8uqUVLtXsKHUQuW5ra0lt-FklrA40qq6_J04yY0nPjlfKG1uPerclUX2gf6axkIioJadYxzOG3cPZJLRcZ2_vHPdipZWvQdICAL2zRnqnOUCGjfq4Q8aMdmA7H6z7xK7W9MEKnIiEALokmtChLtr-s6hDFko17M7xihPpNlfGN7N8D___wn55epkLMtS2eFF3JPlj_SjpFIGXYK15PJFta-BmPqCFvEwXlZEYfEf8uYOpAvCEdBn3NOMoB-P28owOJ7ZeBQf5VMFi3J5_RV2fE_XDR2LTD469Qq0y3946LQ',
    'WORKSPACE_PATH': os.path.expanduser('~/reverseengineerAPI/reverse_engineer_API/workspace'),
    'SHAREPOINT': {
        'SITE_URL': 'deeplydigital.sharepoint.com:/sites/OSPIntegrationTestingSite',  # Updated to new site
        'DRIVE_PATH': 'Documents',  # Simplified path for testing
        'FILE_NAME': 'Aerial_Status_Tracker.xlsx'  # Updated file name
    }
}

# Email configuration
EMAIL_CONFIG = {
    'client_id': os.getenv('AZURE_CLIENT_ID'),
    'client_secret': os.getenv('AZURE_CLIENT_SECRET'),
    'tenant_id': os.getenv('AZURE_TENANT_ID'),
    'user_email': 'brandan.lewis@deeplydigital.com',
    'default_recipients': ['brandan.lewis@deeplydigital.com']
}

# Function to get list of jobs from KatapultPro API
def getJobList():
    URL_PATH = '/api/v2/jobs'
    headers = {}
    all_jobs = []

    for attempt in range(5):  # Consider using exponential backoff to avoid overwhelming the server
        conn = None
        try:
            conn = http.client.HTTPSConnection("katapultpro.com", timeout=10)  # Timeout value could be made configurable
            conn.request("GET", f"{URL_PATH}?api_key={CONFIG['API_KEY']}", headers=headers)
            res = conn.getresponse()
            data = res.read().decode("utf-8")
            jobs_dict = json.loads(data)

            if not isinstance(jobs_dict, dict):
                raise TypeError(f"Expected a dictionary but received {type(jobs_dict)}: {jobs_dict}")

            # Retrieve all jobs without filtering for status
            all_jobs = [
                {'id': job_id, 'name': job_details.get('name'), 'status': job_details.get('status')}
                for job_id, job_details in jobs_dict.items()
            ]
            logging.info(f"Retrieved {len(all_jobs)} jobs")
            break

        except (socket.error, OSError) as e:
            print(f"Socket error: {e}. Retrying...")
            time.sleep(5)
        except Exception as e:
            print(f"Failed to retrieve job list: {e}")
            break
        finally:
            if conn:
                conn.close()

    return all_jobs

# Function to get job data from KatapultPro API
def getJobData(job_id):
    URL_PATH = f'/api/v2/jobs/{job_id}'
    headers = {}
    job_data = None  # Initialize job_data to None
    max_retries = 5
    base_timeout = 60  # Increased base timeout to 60 seconds

    for attempt in range(max_retries):
        conn = None
        current_timeout = base_timeout * (2 ** attempt)  # Exponential backoff for timeout
        wait_time = 5 * (2 ** attempt)  # Exponential backoff for wait time between retries
        
        try:
            print(f"Attempt {attempt + 1}/{max_retries} for job {job_id} (timeout: {current_timeout}s)")
            conn = http.client.HTTPSConnection("katapultpro.com", timeout=current_timeout)
            conn.request("GET", f"{URL_PATH}?api_key={CONFIG['API_KEY']}", headers=headers)
            res = conn.getresponse()
            
            # Check response status
            if res.status != 200:
                print(f"Received status code {res.status} for job {job_id}")
                if res.status == 429:  # Rate limit exceeded
                    print(f"Rate limit exceeded. Waiting {wait_time} seconds before retry...")
                    time.sleep(wait_time)
                    continue
                elif res.status >= 500:  # Server error
                    print(f"Server error. Waiting {wait_time} seconds before retry...")
                    time.sleep(wait_time)
                    continue
            
            data = res.read().decode("utf-8")
            job_data = json.loads(data)

            if "error" in job_data:
                if job_data["error"] == "RATE LIMIT EXCEEDED":
                    print(f"Rate limit exceeded. Waiting {wait_time} seconds before retry...")
                    time.sleep(wait_time)
                    continue
                else:
                    print(f"API error: {job_data['error']}")
                    time.sleep(wait_time)
                    continue

            # Save job data to a file if testing a specific job
            if TEST_ONLY_SPECIFIC_JOB:
                workspace_path = CONFIG['WORKSPACE_PATH']
                file_path = os.path.join(workspace_path, f"test_job_{job_id.replace('/', '_')}.json")
                with open(file_path, 'w') as f:
                    json.dump(job_data, f, indent=2)
                print(f"Job data saved to: {file_path}")
            
            print(f"Successfully retrieved data for job {job_id}")
            return job_data

        except json.JSONDecodeError:
            print(f"Failed to decode JSON for job {job_id}")
            time.sleep(wait_time)
        except (socket.error, OSError) as e:
            print(f"Socket error while retrieving job data for {job_id}: {e}. Waiting {wait_time} seconds before retry...")
            time.sleep(wait_time)
        except Exception as e:
            print(f"Error retrieving job data for {job_id}: {e}")
            time.sleep(wait_time)
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass
            time.sleep(1)  # Small delay between attempts

    print(f"Failed to retrieve job data for job ID {job_id} after {max_retries} attempts.")
    if job_data:  # Only print job_data if it exists
        print(f"Last received Job Data for {job_id}:\n{json.dumps(job_data, indent=2)}")
    return None  # Return None if all attempts fail


# Extract nodes (poles, anchors, etc.) from job data
def extractNodes(job_data, job_name, job_id, user_map):
    nodes = job_data.get("nodes", {})
    if not nodes:
        print("No nodes found.")
        return []

    # Analysis counters (keep existing counters)
    node_type_counts = {}
    attachment_height_counts = {}
    pole_class_counts = {}
    pole_height_counts = {}
    
    # Debug counters
    total_nodes = len(nodes)
    processed_nodes = 0
    skipped_nodes = 0
    nodes_with_height = 0
    nodes_without_height = 0
    
    print(f"\nStarting node analysis for {total_nodes} total nodes...")
    print("------------------------")
    
    photo_data = job_data.get('photos', {})
    trace_data_all = job_data.get('traces', {}).get('trace_data', {})
    node_points = []

    # Extract job status and conversation from job data
    metadata = job_data.get('metadata', {})
    job_status = metadata.get('job_status', "Unknown")
    conversation = metadata.get('conversation', "")

    for node_id, node_data in nodes.items():
        try:
            attributes = node_data.get('attributes', {})
            
            # Extract editor tracking information
            editors_history = {}  # Changed to dict to track latest edit per editor
            last_editor = "Unknown"
            last_edit_time = None
            
            # Check photos associated with the node for editor information
            node_photos = node_data.get('photos', {})
            for photo_id in node_photos:
                if photo_id in photo_data:
                    photo_editors = photo_data[photo_id].get('photofirst_data', {}).get('_editors', {})
                    if photo_editors:
                        # Process all editors for this photo
                        for editor_id, edit_time in photo_editors.items():
                            editor_name = user_map.get(editor_id, "Unknown User")
                            # Only update if this is the most recent edit for this editor
                            if editor_name not in editors_history or edit_time > editors_history[editor_name]['raw_timestamp']:
                                # Convert timestamp to MST (UTC-7)
                                edit_dt = datetime.fromtimestamp(edit_time/1000)
                                # Subtract 7 hours for MST
                                edit_dt_mst = edit_dt.replace(hour=(edit_dt.hour - 7) % 24)
                                formatted_time = edit_dt_mst.strftime('%I:%M %p MST')
                                formatted_date = edit_dt_mst.strftime('%Y-%m-%d')
                                
                                editors_history[editor_name] = {
                                    'editor': editor_name,
                                    'timestamp': f"{formatted_date} {formatted_time}",
                                    'raw_timestamp': edit_time
                                }
            
            # Convert editors_history dict to list and sort by timestamp
            editors_list = list(editors_history.values())
            if editors_list:
                editors_list.sort(key=lambda x: x['raw_timestamp'], reverse=True)
                last_editor = editors_list[0]['editor']
                last_edit_time = editors_list[0]['timestamp']
                
                # Print detailed editor history for debugging
                print(f"\nEditor history for node {node_id}:")
                for edit in editors_list:
                    print(f"  Editor: {edit['editor']}")
                    print(f"  Last Edit: {edit['timestamp']}")
                print("------------------------")
            
            # Check if node is a pole
            node_type = 'Unknown'
            for type_source in ['node_type', 'pole_type']:
                for source_type in ['-Imported', 'button_added', 'value', 'auto_calced']:
                    type_value = attributes.get(type_source, {}).get(source_type)
                    if type_value:
                        node_type = type_value
                        break
                if node_type != 'Unknown':
                    break
            
            # Count node types
            if node_type not in node_type_counts:
                node_type_counts[node_type] = 0
            node_type_counts[node_type] += 1
            
            # Only process poles (exclude anchors, references, etc.)
            if node_type == 'pole':
                processed_nodes += 1
                latitude = node_data.get('latitude')
                longitude = node_data.get('longitude')

                if latitude is None or longitude is None:
                    print(f"Warning: Missing coordinates for pole {node_id}")
                    skipped_nodes += 1
                    continue

                # Extract MR status
                mr_status = "Unknown"
                if 'proposed_pole_spec' in attributes:
                    mr_status = "PCO Required"
                else:
                    mr_state = attributes.get('mr_state', {}).get('auto_calced', "Unknown")
                    warning_present = 'warning' in attributes
                    if mr_state == "No MR" and not warning_present:
                        mr_status = "No MR"
                    elif mr_state == "MR Resolved" and not warning_present:
                        mr_status = "Comm MR"
                    elif mr_state == "MR Resolved" and warning_present:
                        mr_status = "Electric MR"

                # Extract pole attributes
                company = attributes.get('pole_tag', {}).get('-Imported', {}).get('company', "Unknown")
                fldcompl_value = attributes.get('field_completed', {}).get('value', "Unknown")
                fldcompl = 'yes' if fldcompl_value == 1 else 'no' if fldcompl_value == 2 else 'Unknown'
                
                # Extract and count pole class
                pole_class = attributes.get('pole_class', {}).get('-Imported', "Unknown")
                if pole_class not in pole_class_counts:
                    pole_class_counts[pole_class] = 0
                pole_class_counts[pole_class] += 1
                
                # Extract and count pole height
                pole_height = attributes.get('pole_height', {}).get('-Imported', "Unknown")
                if pole_height not in pole_height_counts:
                    pole_height_counts[pole_height] = 0
                pole_height_counts[pole_height] += 1
                
                # Extract tag and scid
                tag = attributes.get('pole_tag', {}).get('-Imported', {}).get('tagtext', "Unknown")
                scid = attributes.get('scid', {}).get('auto_button', "Unknown")

                # Extract POA height with detailed analysis
                poa_height = ""
                photos = node_data.get('photos', {})
                main_photo_id = next(
                    (photo_id for photo_id, photo_info in photos.items() if photo_info.get('association') == 'main'), None)

                if main_photo_id and main_photo_id in photo_data:
                    # Check wire data
                    photofirst_data = photo_data[main_photo_id].get('photofirst_data', {}).get('wire', {})
                    for wire_info in photofirst_data.values():
                        trace_id = wire_info.get('_trace')
                        trace_data = trace_data_all.get(trace_id, {})

                        if (trace_data.get('company') == 'Clearnetworx' and
                                trace_data.get('proposed', False) and
                                trace_data.get('_trace_type') == 'cable' and
                                trace_data.get('cable_type') == 'Fiber Optic Com'):

                            poa_height = wire_info.get('_measured_height')
                            if poa_height is not None:
                                feet = int(poa_height // 12)
                                inches = int(poa_height % 12)
                                poa_height = f"{feet}' {inches}\""
                                nodes_with_height += 1
                                
                                # Count attachment heights
                                height_key = f"{feet}'{inches}\""
                                if height_key not in attachment_height_counts:
                                    attachment_height_counts[height_key] = 0
                                attachment_height_counts[height_key] += 1
                                break

                    # Check guying data if no POA height found
                    if not poa_height:
                        guying_data = photo_data[main_photo_id].get('photofirst_data', {}).get('guying', {})
                        for wire_info in guying_data.values():
                            trace_id = wire_info.get('_trace')
                            trace_data = trace_data_all.get(trace_id, {})

                            if (trace_data.get('company') == 'Clearnetworx' and
                                    trace_data.get('proposed', False) and
                                    trace_data.get('_trace_type') == 'down_guy'):

                                poa_height = wire_info.get('_measured_height')
                                if poa_height is not None:
                                    feet = int(poa_height // 12)
                                    inches = int(poa_height % 12)
                                    poa_height = f"{feet}' {inches}\""
                                    nodes_with_height += 1
                                    
                                    # Count attachment heights
                                    height_key = f"{feet}'{inches}\""
                                    if height_key not in attachment_height_counts:
                                        attachment_height_counts[height_key] = 0
                                    attachment_height_counts[height_key] += 1
                                    break
                
                if not poa_height:
                    nodes_without_height += 1

                node_points.append({
                    "id": node_id,
                    "lat": latitude,
                    "lng": longitude,
                    "jobname": job_name,
                    "job_status": job_status,
                    "MR_statu": mr_status,
                    "company": company,
                    "fldcompl": fldcompl,
                    "pole_class": pole_class,
                    "tag": tag,
                    "scid": scid,
                    "POA_Height": poa_height,
                    "conversation": conversation,
                    "last_editor": last_editor,
                    "last_edit": last_edit_time
                })
        except Exception as e:
            print(f"Error processing node {node_id}: {str(e)}")
            skipped_nodes += 1

    # Print analysis results
    print("\nNode Processing Summary:")
    print(f"Total nodes found: {total_nodes}")
    print(f"Poles processed: {processed_nodes}")
    print(f"Skipped nodes: {skipped_nodes}")
    print(f"Poles with height: {nodes_with_height}")
    print(f"Poles without height: {nodes_without_height}")
    
    print("\nNode Type Distribution:")
    for ntype, count in sorted(node_type_counts.items()):
        print(f"{ntype}: {count}")
    
    print("\nPole Class Distribution:")
    for pclass, count in sorted(pole_class_counts.items()):
        print(f"{pclass}: {count}")
    
    print("\nPole Height Distribution:")
    for pheight, count in sorted(pole_height_counts.items()):
        print(f"{pheight}: {count}")
    
    print("\nAttachment Height Distribution:")
    for height, count in sorted(attachment_height_counts.items()):
        print(f"{height}: {count}")
    
    print("------------------------\n")

    return node_points

def extractAnchors(job_data, job_name, job_id):
    """Extract anchor points from job data, with detailed type analysis."""
    anchors = job_data.get("nodes", {})
    anchor_points = []
    
    # Analysis counters
    node_type_counts = {}
    anchor_spec_counts = {}
    anchor_status = {}
    
    print("\nDetailed Anchor Analysis:")
    print("------------------------")
    
    for node_id, node_data in anchors.items():
        attributes = node_data.get("attributes", {})
        
        # Check all possible node type fields
        node_type = "Unknown"
        for type_field in ["node_type", "anchor_type"]:
            for source in ["button_added", "-Imported", "value", "auto_calced"]:
                type_value = attributes.get(type_field, {}).get(source)
                if type_value and "anchor" in str(type_value).lower():
                    node_type = type_value
                    break
            if node_type != "Unknown":
                break
                
        # Count node types
        if node_type not in node_type_counts:
            node_type_counts[node_type] = 0
        node_type_counts[node_type] += 1
        
        # Only process if it's an anchor
        if "anchor" in str(node_type).lower():
            latitude = node_data.get("latitude")
            longitude = node_data.get("longitude")
            
            # Get anchor specification with detailed logging
            anchor_spec = "Unknown"
            print(f"\nAnalyzing anchor {node_id} (Type: {node_type}):")
            
            # Check anchor_spec field for both multi_added and button_added
            anchor_spec_data = attributes.get("anchor_spec", {})
            if anchor_spec_data.get("multi_added"):
                anchor_spec = anchor_spec_data.get("multi_added")
                print(f"  Found spec in multi_added: {anchor_spec}")
            elif anchor_spec_data.get("button_added"):
                anchor_spec = anchor_spec_data.get("button_added")
                print(f"  Found spec in button_added: {anchor_spec}")
            
            # Count anchor specs
            if anchor_spec not in anchor_spec_counts:
                anchor_spec_counts[anchor_spec] = 0
            anchor_spec_counts[anchor_spec] += 1
            
            # Track anchor status (new vs existing)
            status = "new" if "new" in str(node_type).lower() else "existing" if "existing" in str(node_type).lower() else "unknown"
            if status not in anchor_status:
                anchor_status[status] = 0
            anchor_status[status] += 1
            
            # Append anchor information
            anchor_points.append({
                "longitude": longitude,
                "latitude": latitude,
                "anchor_spec": anchor_spec,
                "anchor_type": node_type,
                "job_id": job_id
            })
    
    # Print summary analysis
    print("\nSummary Analysis:")
    print("------------------------")
    print("Node Types Found:")
    for ntype, count in sorted(node_type_counts.items()):
        print(f"{ntype}: {count}")
    
    print("\nAnchor Specifications:")
    for spec, count in sorted(anchor_spec_counts.items()):
        print(f"{spec}: {count}")
    
    print("\nAnchor Status:")
    for status, count in sorted(anchor_status.items()):
        print(f"{status}: {count}")
    print("------------------------")

    return anchor_points

# Extract connections (lines, cables, etc.) from job data
def extractConnections(connections, nodes):
    # Analysis counters
    connection_type_counts = {}
    connection_height_counts = {}
    
    # Debug counters
    total_connections = len(connections)
    connections_with_height = 0
    connections_without_height = 0
    connections_by_type = {}
    
    print(f"\nStarting connection analysis for {total_connections} total connections...")
    print("------------------------")
    
    # First analyze connection types and heights
    for connection_id, connection_data in connections.items():
        try:
            # Check both value and button_added for connection type
            connection_type = 'Unknown'
            attributes = connection_data.get('attributes', {}).get('connection_type', {})
            
            # First check button_added
            if attributes.get('button_added'):
                connection_type = attributes.get('button_added')
            # Then check value if still unknown
            elif attributes.get('value'):
                connection_type = attributes.get('value')
                
            if connection_type not in connection_type_counts:
                connection_type_counts[connection_type] = 0
            connection_type_counts[connection_type] += 1
            
            # Track connections by type for debugging
            if connection_type not in connections_by_type:
                connections_by_type[connection_type] = []
            connections_by_type[connection_type].append(connection_id)
            
            # Extract attachment height if available
            sections = connection_data.get('sections', {}).get('midpoint_section', {})
            if 'attachment_height' in sections:
                height = sections['attachment_height']
                if height is not None:
                    feet = int(height // 12)
                    inches = int(height % 12)
                    height_key = f"{feet}'{inches}\""
                    if height_key not in connection_height_counts:
                        connection_height_counts[height_key] = 0
                    connection_height_counts[height_key] += 1
                    connections_with_height += 1
                else:
                    connections_without_height += 1
            else:
                connections_without_height += 1
                
        except Exception as e:
            print(f"Error analyzing connection {connection_id}: {str(e)}")
    
    print("\nConnection Type Distribution:")
    for conn_type, count in sorted(connection_type_counts.items()):
        print(f"{conn_type}: {count}")
        print(f"  Sample IDs: {connections_by_type[conn_type][:3]}")  # Show first 3 IDs for each type
        
    print("\nConnection Height Distribution:")
    for height, count in sorted(connection_height_counts.items()):
        print(f"{height}: {count}")
    
    print("\nConnection Height Summary:")
    print(f"Connections with height: {connections_with_height}")
    print(f"Connections without height: {connections_without_height}")
    print("------------------------\n")

    valid_connections = []
    processed_count = 0
    skipped_count = 0
    
    for connection_id, connection_data in connections.items():
        try:
            node_id_1 = connection_data.get('node_id_1')
            node_id_2 = connection_data.get('node_id_2')
            
            if node_id_1 not in nodes or node_id_2 not in nodes:
                print(f"Warning: Missing node(s) for connection {connection_id} (nodes: {node_id_1}, {node_id_2})")
                skipped_count += 1
                continue
                
            start_node = nodes[node_id_1]
            end_node = nodes[node_id_2]
            
            start_lat = start_node.get('latitude')
            start_lon = start_node.get('longitude')
            end_lat = end_node.get('latitude')
            end_lon = end_node.get('longitude')
            
            if any(coord is None for coord in [start_lat, start_lon, end_lat, end_lon]):
                print(f"Warning: Missing coordinates for connection {connection_id}")
                skipped_count += 1
                continue
            
            line = LineString([(start_lon, start_lat), (end_lon, end_lat)])
            
            # Get connection type from both possible sources
            attributes = connection_data.get('attributes', {}).get('connection_type', {})
            connection_type = 'Unknown'
            if attributes.get('button_added'):
                connection_type = attributes.get('button_added')
            elif attributes.get('value'):
                connection_type = attributes.get('value')
            
            # Get attachment height if available
            attachment_height = None
            sections = connection_data.get('sections', {}).get('midpoint_section', {})
            if 'attachment_height' in sections:
                height = sections['attachment_height']
                if height is not None:
                    feet = int(height // 12)
                    inches = int(height % 12)
                    attachment_height = f"{feet}' {inches}\""
            
            feature = {
                'type': 'Feature',
                'geometry': mapping(line),
                'properties': {
                    'connection_id': connection_id,
                    'connection_type': connection_type,
                    'attachment_height': attachment_height,
                    'StartX': start_lon,
                    'StartY': start_lat,
                    'EndX': end_lon,
                    'EndY': end_lat,
                    'node_id_1': node_id_1,
                    'node_id_2': node_id_2
                }
            }
            
            valid_connections.append(feature)
            processed_count += 1
            
        except Exception as e:
            print(f"Error processing connection {connection_id}: {str(e)}")
            skipped_count += 1
            continue
    
    print(f"\nConnection processing summary:")
    print(f"Total connections: {total_connections}")
    print(f"Successfully processed: {processed_count}")
    print(f"Skipped: {skipped_count}")
    
    return valid_connections



def savePointsToShapefile(points, filename):
    workspace_path = CONFIG['WORKSPACE_PATH']
    file_path = os.path.join(workspace_path, filename.replace('.shp', '.gpkg'))
    geometries = [Point(point["lng"], point["lat"]) for point in points]

    gdf = gpd.GeoDataFrame(points, geometry=geometries, crs="EPSG:4326")

    # Rename columns
    gdf.rename(columns={
        'company': 'utility',
        'tag': 'pole tag',
        'fldcompl': 'collected',
        'jobname': 'jobname',
        'job_status': 'job_status',
        'MR_statu': 'mr_status',
        'pole_spec': 'pole_spec',
        'POA_Height': 'att_ht',
        'lat': 'latitude',
        'lng': 'longitude'
    }, inplace=True)

    # Remove unwanted columns, ignore if they don't exist
    gdf.drop(columns=['pole_class', 'pole_height', 'id'], errors='ignore', inplace=True)

    # Save to file
    try:
        gdf.to_file(file_path, driver="GPKG")  # Switched to GeoPackage for better flexibility
        print(f"GeoPackage successfully saved to: {file_path}")
    except Exception as e:
        print(f"Error saving GeoPackage: {e}")


# Function to save line connections to a GeoPackage
def saveAnchorsToGeoPackage(anchor_points, filename):
    workspace_path = CONFIG['WORKSPACE_PATH']
    file_path = os.path.join(workspace_path, filename.replace('.shp', '.gpkg'))
    geometries = [Point(anchor["longitude"], anchor["latitude"]) for anchor in anchor_points]

    gdf = gpd.GeoDataFrame(anchor_points, geometry=geometries, crs="EPSG:4326")

    # Rename columns
    gdf.rename(columns={
        'longitude': 'longitude',
        'latitude': 'latitude',
        'anchor_spec': 'anchor_spec'
    }, inplace=True)

    # Save to file
    try:
        gdf.to_file(file_path, layer='anchors', driver="GPKG")
        print(f"Anchors GeoPackage successfully saved to: {file_path}")
    except Exception as e:
        print(f"Error saving anchors GeoPackage: {e}")
def saveLineShapefile(line_connections, filename):
    workspace_path = CONFIG['WORKSPACE_PATH']
    file_path = os.path.join(workspace_path, filename.replace('.shp', '.gpkg'))
    geometries = [
        LineString([(line["StartX"], line["StartY"]), (line["EndX"], line["EndY"])])
        for line in line_connections
    ]

    gdf = gpd.GeoDataFrame(line_connections, geometry=geometries, crs="EPSG:4326")
    gdf.drop(columns=['StartX', 'StartY', 'EndX', 'EndY', 'job_id'], errors='ignore', inplace=True)
    try:
        gdf.to_file(file_path, driver="GPKG")  # Switched to GeoPackage for better flexibility
        print(f"Line GeoPackage successfully saved to: {file_path}")
    except Exception as e:
        print(f"Error saving line GeoPackage: {e}")

# Function to save nodes to a GeoPackage
def saveMasterGeoPackage(all_nodes, all_connections, all_anchors, filename):
    workspace_path = CONFIG['WORKSPACE_PATH']
    file_path = os.path.join(workspace_path, filename)

    # Save nodes as point layer
    if all_nodes:
        try:
            # Create point geometries for nodes
            geometries = [Point(node["lng"], node["lat"]) for node in all_nodes]
            gdf_nodes = gpd.GeoDataFrame(all_nodes, geometry=geometries, crs="EPSG:4326")
            
            # Drop lat/lng columns as they're now in the geometry
            gdf_nodes.drop(columns=['lat', 'lng'], errors='ignore', inplace=True)
            
            # Save to GeoPackage
            gdf_nodes.to_file(file_path, layer='nodes', driver="GPKG")
            print(f"Nodes layer successfully saved to: {file_path}")
            
        except Exception as e:
            print(f"Error saving nodes layer to GeoPackage: {e}")

    # Save connections as line layer
    if all_connections:
        try:
            valid_connections = []
            line_geometries = []
            
            print(f"Processing {len(all_connections)} connections...")
            for connection in all_connections:
                try:
                    # Get coordinates from the connection properties
                    properties = connection.get('properties', {})
                    start_x = properties.get('StartX')
                    start_y = properties.get('StartY')
                    end_x = properties.get('EndX')
                    end_y = properties.get('EndY')

                    if any(coord is None for coord in [start_x, start_y, end_x, end_y]):
                        print(f"Missing coordinates for connection between nodes {properties.get('node_id_1')} and {properties.get('node_id_2')}")
                        continue

                    # Create LineString geometry
                    line_geom = LineString([(start_x, start_y), (end_x, end_y)])
                    
                    # Use the existing properties
                    valid_connections.append(properties)
                    line_geometries.append(line_geom)
                        
                except Exception as e:
                    print(f"Error processing line: {str(e)}")
                    continue
            
            print(f"Found {len(valid_connections)} valid connections out of {len(all_connections)} total connections")
            
            if valid_connections and line_geometries:
                gdf_lines = gpd.GeoDataFrame(valid_connections, geometry=line_geometries, crs="EPSG:4326")
                
                if not gdf_lines.empty:
                    gdf_lines.to_file(file_path, layer='connections', driver="GPKG", mode='a')
                    print("Connections layer successfully saved to GeoPackage")
            else:
                print("No valid connections found")
                
        except Exception as e:
            print(f"Error saving connections layer to GeoPackage: {e}")

    # Save anchors as point layer
    if all_anchors:
        try:
            # Create point geometries for anchors
            geometries = [Point(anchor["longitude"], anchor["latitude"]) for anchor in all_anchors]
            gdf_anchors = gpd.GeoDataFrame(all_anchors, geometry=geometries, crs="EPSG:4326")
            
            # Drop lat/lng columns as they're now in the geometry
            gdf_anchors.drop(columns=['latitude', 'longitude'], errors='ignore', inplace=True)
            
            # Save to GeoPackage
            gdf_anchors.to_file(file_path, layer='anchors', driver="GPKG", mode='a')
            print("Anchors layer successfully saved to GeoPackage")
            
        except Exception as e:
            print(f"Error saving anchors layer to GeoPackage: {e}")

    print("Master GeoPackage saved successfully")

def update_sharepoint_spreadsheet(df, site_url=None, drive_path=None):
    """
    Update the spreadsheet in SharePoint with new data, supporting co-authoring
    """
    try:
        print("\nUpdating SharePoint spreadsheet...")
        
        # Format the data for SharePoint
        headers = list(df.columns)
        formatted_data = [headers]  # First row is headers
        formatted_data.extend(df.values.tolist())  # Add all data rows
        
        # Use configured paths or fallback to parameters
        site_url = site_url or CONFIG['SHAREPOINT']['SITE_URL']
        drive_path = drive_path or CONFIG['SHAREPOINT']['DRIVE_PATH']
        file_name = CONFIG['SHAREPOINT']['FILE_NAME']
        file_path = f"{drive_path}/{file_name}"
        
        print(f"Using SharePoint path: {site_url}")
        print(f"Using file path: {file_path}")
        
        # Initialize the Graph client
        graph_client = initialize_graph_client()
        
        if not graph_client:
            print("Failed to initialize Graph client")
            return False
            
        # Get the site ID
        print("Getting site ID...")
        site_response = graph_client.get(f"sites/{site_url}")
        if site_response.status_code != 200:
            print(f"Failed to get site. Status code: {site_response.status_code}")
            print(f"Response: {site_response.text}")
            return False
            
        site_id = site_response.json()['id']
        print(f"Successfully got site ID: {site_id}")
        
        # Get the drive ID
        print("Getting drive ID...")
        drives_response = graph_client.get(f"sites/{site_id}/drives")
        if drives_response.status_code != 200:
            print(f"Failed to get drives. Status code: {drives_response.status_code}")
            print(f"Response: {drives_response.text}")
            return False
            
        # Find the Documents drive
        documents_drive = None
        for drive in drives_response.json()['value']:
            if drive['name'] == 'Documents':
                documents_drive = drive
                break
                
        if not documents_drive:
            print("Could not find Documents drive")
            return False
            
        drive_id = documents_drive['id']
        print(f"Successfully got drive ID: {drive_id}")
        
        # Get the file
        print(f"Checking for existing file at: {file_path}")
        file_response = graph_client.get(f"sites/{site_id}/drives/{drive_id}/root:/{file_path}")
        
        if file_response.status_code == 200:
            print("File exists, updating...")
            file_id = file_response.json()['id']
            
            # Create a workbook session
            session_response = graph_client.post(
                f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/createSession",
                json={"persistChanges": True}
            )
            
            if session_response.status_code != 201:
                print(f"Failed to create workbook session. Status code: {session_response.status_code}")
                print(f"Response: {session_response.text}")
                return False
            
            session_id = session_response.json()['id']
            print("Successfully created workbook session")
            
            try:
                # Get the worksheet
                worksheet_response = graph_client.get(
                    f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report",
                    headers={"workbook-session-id": session_id}
                )
                
                if worksheet_response.status_code != 200:
                    # Try to add the worksheet if it doesn't exist
                    worksheet_response = graph_client.post(
                        f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets",
                        headers={"workbook-session-id": session_id},
                        json={"name": "Aerial Status Report"}
                    )
                    
                    if worksheet_response.status_code != 201:
                        print(f"Failed to create worksheet. Status code: {worksheet_response.status_code}")
                        print(f"Response: {worksheet_response.text}")
                        return False

                # Update the title and date with formatting and merging
                # First, merge cells for title (A1:L1)
                graph_client.post(
                    f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report/range(address='A1:L1')/merge",
                    headers={"workbook-session-id": session_id}
                )

                # Then update title with formatting
                graph_client.patch(
                    f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report/range(address='A1:L1')",
                    headers={"workbook-session-id": session_id},
                    json={
                        "values": [["Aerial Status Report"]],
                        "format": {
                            "font": {"bold": True, "size": 18},
                            "horizontalAlignment": "center",
                            "verticalAlignment": "center"
                        }
                    }
                )

                # Update timestamp
                current_date = datetime.now().strftime('%m/%d/%Y %I:%M %p')
                print(f"Updating timestamp to: {current_date}")
                
                # Update timestamp with correct array dimensions
                timestamp_response = graph_client.patch(
                    f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report/range(address='A2:L2')",
                    headers={"workbook-session-id": session_id},
                    json={
                        "values": [[current_date] + [""] * 11],  # One row with 12 columns
                        "numberFormat": [["@"] * 12]  # Format for all 12 columns
                    }
                )
                
                if timestamp_response.status_code != 200:
                    print(f"Failed to update timestamp. Status code: {timestamp_response.status_code}")
                    print(f"Response: {timestamp_response.text}")
                    return False

                # Commit the changes immediately after timestamp update
                commit_response = graph_client.post(
                    f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/functions/saveWorkbook",
                    headers={"workbook-session-id": session_id}
                )
                
                if commit_response.status_code != 200:
                    print(f"Failed to commit changes. Status code: {commit_response.status_code}")
                    print(f"Response: {commit_response.text}")
                    return False

                # Set column widths
                column_widths = {
                    "A": 44,  # Job Name
                    "B": 23.71,  # Job Status
                    "C": 30,  # Last Editor
                    "D": 20,  # Last Edit
                    "E": 15,  # Utility
                    "F": 10,  # Field %
                    "G": 10,  # Trace %
                    "H": 10,  # No MR
                    "I": 10,  # Comm MR
                    "J": 12,  # Electric MR
                    "K": 12,  # PCO Required
                    "L": 12,  # Pole Count
                }

                # Apply column widths
                for col_letter, width in column_widths.items():
                    graph_client.patch(
                        f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report/column('{col_letter}')",
                        headers={"workbook-session-id": session_id},
                        json={"columnWidth": width}
                    )

                # Set row heights
                row_heights = {
                    1: 30,  # Title row
                    2: 20,  # Date row
                    3: 25,  # Column headers row
                }

                for row_num, height in row_heights.items():
                    graph_client.patch(
                        f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report/row({row_num})",
                        headers={"workbook-session-id": session_id},
                        json={"height": height}
                    )

                # Define header colors and apply formatting
                header_colors = {
                    "A": "CCFFCC",  # Job Name (Green)
                    "B": "CCFFCC",  # Job Status (Green)
                    "C": "CCFFCC",  # Last Editor (Green)
                    "D": "CCFFCC",  # Last Edit (Green)
                    "E": "CCFFCC",  # Utility (Green)
                    "F": "CCFFCC",  # Field % (Green)
                    "G": "CCFFCC",  # Trace % (Green)
                    "H": "D9D9D9",  # No MR (Gray)
                    "I": "FFFF00",  # Comm MR (Yellow)
                    "J": "FFC000",  # Electric MR (Orange)
                    "K": "FF0000",  # PCO Required (Red)
                    "L": "CCFFCC",  # Pole Count (Green)
                }

                # Apply header formatting
                for col_letter, color in header_colors.items():
                    header_range = f"{col_letter}3"
                    graph_client.patch(
                        f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report/range(address='{header_range}')",
                        headers={"workbook-session-id": session_id},
                        json={
                            "format": {
                                "fill": {"color": color},
                                "font": {"bold": True},
                                "horizontalAlignment": "center",
                                "verticalAlignment": "center",
                                "borders": {
                                    "allBorders": {
                                        "style": "thin",
                                        "color": "#000000"
                                    }
                                }
                            }
                        }
                    )

                # Update the data rows with center alignment and borders
                data_range = f"A4:L{len(formatted_data) + 2}"  # Start from row 4 to preserve title, timestamp, and headers
                graph_client.patch(
                    f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/worksheets/Aerial%20Status%20Report/range(address='{data_range}')",
                    headers={"workbook-session-id": session_id},
                    json={
                        "values": formatted_data[1:],  # Skip the header row
                        "format": {
                            "horizontalAlignment": "center",
                            "verticalAlignment": "center",
                            "borders": {
                                "allBorders": {
                                    "style": "thin",
                                    "color": "#000000"
                                }
                            }
                        }
                    }
                )

                print("Successfully updated data and formatting")
                return True

            finally:
                # Close the session
                try:
                    graph_client.post(
                        f"sites/{site_id}/drives/{drive_id}/items/{file_id}/workbook/closeSession",
                        headers={"workbook-session-id": session_id}
                    )
                except Exception as e:
                    print(f"Error closing session: {str(e)}")

        else:
            print("File doesn't exist, creating new file...")
            # Create new file logic remains the same...

    except Exception as e:
        print(f"Error updating SharePoint spreadsheet: {str(e)}")
        return False

# Modify create_report function to include SharePoint update
def create_report(jobs_summary):
    report_data = []

    for job in jobs_summary:
        job_name = job['job_name']
        job_status = job.get('job_status', 'Unknown').strip()
        mr_status_counts = job['mr_status_counts']
        pole_count = sum(mr_status_counts.values())
        
        # Get the fields from job summary
        field_complete_pct = job.get('field_complete_pct', 0)
        trace_complete_pct = job.get('trace_complete_pct', 0)
        utility = job.get('utility', 'Unknown')
        most_recent_editor = job.get('most_recent_editor', 'Unknown')
        last_edit_time = job.get('last_edit_time', 'Unknown')

        report_data.append({
            'Job Name': job_name,
            'Job Status': job_status,
            'Last Editor': most_recent_editor,
            'Last Edit': last_edit_time,
            'Utility': utility,
            'Field %': f"{field_complete_pct:.1f}%",
            'Trace %': f"{trace_complete_pct:.1f}%",
            'No MR': mr_status_counts.get('No MR', 0),
            'Comm MR': mr_status_counts.get('Comm MR', 0),
            'Electric MR': mr_status_counts.get('Electric MR', 0),
            'PCO Required': mr_status_counts.get('PCO Required', 0),
            'Pole Count': pole_count
        })

    # Create a DataFrame from the report data
    df_report = pd.DataFrame(report_data)

    # Sort the DataFrame by the "Job Status" column alphabetically
    df_report = df_report.sort_values(by='Job Status')

    # Ensure the directory exists
    workspace_dir = CONFIG['WORKSPACE_PATH']
    if not os.path.exists(workspace_dir):
        try:
            os.makedirs(workspace_dir)
            print(f"Workspace directory created: {workspace_dir}")
        except Exception as e:
            print(f"Failed to create workspace directory: {e}")
            return None

    # Generate a filename with a timestamp
    timestamp = datetime.now().strftime("%m%d%Y_%I%M")
    report_filename = f"Aerial_Status_Report_{timestamp}.xlsx"
    report_path = os.path.join(workspace_dir, report_filename)

    # Write the report to an Excel file with formatting
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Aerial Status Report"

        # Add merged header with title in the first row
        ws.merge_cells('A1:G1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "Aerial Status Report"
        title_cell.font = Font(size=18, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add the date/time in the second row
        ws.merge_cells('A2:G2')
        date_cell = ws.cell(row=2, column=1)
        date_cell.value = datetime.now().strftime('%-m/%-d/%Y %-I:%M %p')
        date_cell.font = Font(size=12)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set row heights
        ws.row_dimensions[1].height = 30  # Title row
        ws.row_dimensions[2].height = 20  # Date row
        ws.row_dimensions[3].height = 25  # Column headers row

        # Add the column headers with styling in the third row
        column_widths = {
            "Job Name": 44,
            "Job Status": 23.71,
            "Last Editor": 30,
            "Last Edit": 20,
            "Utility": 15,
            "Field %": 10,
            "Trace %": 10,
            "No MR": 10,
            "Comm MR": 10,
            "Electric MR": 12,
            "PCO Required": 12,
            "Pole Count": 12
        }

        header_colors = {
            "Job Name": "CCFFCC",
            "Job Status": "CCFFCC",
            "Last Editor": "CCFFCC",
            "Last Edit": "CCFFCC",
            "Utility": "CCFFCC",
            "Field %": "CCFFCC",
            "Trace %": "CCFFCC",
            "No MR": "D9D9D9",
            "Comm MR": "FFFF00",
            "Electric MR": "FFC000",
            "PCO Required": "FF0000",
            "Pole Count": "CCFFCC"
        }

        for col_num, column_title in enumerate(df_report.columns, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column width
            column_letter = cell.column_letter
            ws.column_dimensions[column_letter].width = column_widths.get(column_title, 13.3)

            # Set header color
            if column_title in header_colors:
                cell.fill = PatternFill(start_color=header_colors[column_title],
                                      end_color=header_colors[column_title],
                                      fill_type="solid")

        # Add the data rows with center alignment
        for r_idx, row in enumerate(dataframe_to_rows(df_report, index=False, header=False), 4):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add borders around all cells
        thin_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(df_report.columns)):
            for cell in row:
                cell.border = thin_border

        # Add Job Status Summary Headers with colors and proper spacing
        status_summary_start_row = 6
        job_statuses = [
            ("Pending Field Collection", "CCFFCC"),
            ("Pending Photo Annotation", "B7DEE8"),
            ("Sent to PE", "CCC0DA"),
            ("Pending EMR", "FFC000"),
            ("Approved for Construction", "9BBB59"),
            ("Hold", "BFBFBF"),
            ("As Built", "FABF8F"),
            ("Delivered", "92D050")
        ]

        # Calculate job status counts
        job_status_counts = {status[0]: 0 for status in job_statuses}
        for job in jobs_summary:
            job_status = job.get('job_status', 'Unknown').strip()
            if job_status in job_status_counts:
                job_status_counts[job_status] += 1

        # Add status headers and counts in two rows, starting from column N (14)
        for idx, (status, color) in enumerate(job_statuses[:4]):
            header_cell = ws.cell(row=status_summary_start_row, column=idx + 14)
            count_cell = ws.cell(row=status_summary_start_row + 1, column=idx + 14)
            
            header_cell.value = status
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            count_cell.value = job_status_counts[status]
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.column_dimensions[header_cell.column_letter].width = 24.14

        # Add second row of statuses
        for idx, (status, color) in enumerate(job_statuses[4:]):
            header_cell = ws.cell(row=status_summary_start_row + 3, column=idx + 14)
            count_cell = ws.cell(row=status_summary_start_row + 4, column=idx + 14)
            
            header_cell.value = status
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            count_cell.value = job_status_counts[status]
            count_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add borders to status summary
        for row in ws.iter_rows(min_row=status_summary_start_row,
                              max_row=status_summary_start_row + 4,
                              min_col=14, max_col=17):
            for cell in row:
                cell.border = thin_border

        # Save the workbook
        wb.save(report_path)
        print(f"Report successfully created: {report_path}")
    except Exception as e:
        print(f"Error creating report: {e}")

    try:
        # After creating the local Excel file
        print("\nUpdating SharePoint spreadsheet...")
        sharepoint_update_success = update_sharepoint_spreadsheet(
            df_report,
            CONFIG['SHAREPOINT']['SITE_URL'],
            CONFIG['SHAREPOINT']['DRIVE_PATH']
        )
        if sharepoint_update_success:
            print("SharePoint spreadsheet updated successfully")
        else:
            print("Failed to update SharePoint spreadsheet")
            
    except Exception as e:
        print(f"Error in SharePoint update: {str(e)}")
    
    return report_path

# Function to send email notification with attachment
def send_email_notification(recipients, report_path):
    """Send email notification with the report attached."""
    print("\nStarting email notification process...")
    print(f"Recipients: {recipients}")
    
    try:
        # Load environment variables
        load_dotenv()
        client_id = os.getenv('AZURE_CLIENT_ID')
        client_secret = os.getenv('AZURE_CLIENT_SECRET')
        tenant_id = os.getenv('AZURE_TENANT_ID')
        user_email = os.getenv('EMAIL_USER')

        # Initialize MSAL client
        authority = f"https://login.microsoftonline.com/{tenant_id}"
        app = msal.ConfidentialClientApplication(
            client_id,
            authority=authority,
            client_credential=client_secret
        )

        # Get access token
        scopes = ['https://graph.microsoft.com/.default']
        result = app.acquire_token_silent(scopes, account=None)
        if not result:
            result = app.acquire_token_for_client(scopes)

        if 'access_token' in result:
            # Prepare email message
            email_msg = {
                'message': {
                    'subject': 'Aerial Status Report Generated',
                    'body': {
                        'contentType': 'Text',
                        'content': 'Please find attached the latest Aerial Status Report.'
                    },
                    'toRecipients': [{'emailAddress': {'address': r}} for r in recipients],
                    'attachments': [{
                        '@odata.type': '#microsoft.graph.fileAttachment',
                        'name': os.path.basename(report_path),
                        'contentBytes': base64.b64encode(open(report_path, 'rb').read()).decode()
                    }]
                }
            }

            # Send email using Microsoft Graph API
            graph_endpoint = 'https://graph.microsoft.com/v1.0'
            headers = {
                'Authorization': f"Bearer {result['access_token']}",
                'Content-Type': 'application/json'
            }
            
            response = requests.post(
                f"{graph_endpoint}/users/{user_email}/sendMail",
                headers=headers,
                data=json.dumps(email_msg)
            )
            
            if response.status_code == 202:
                print("Email sent successfully")
            else:
                print(f"Failed to send email. Status code: {response.status_code}")
                print(f"Response: {response.text}")
        else:
            print(f"Error getting access token: {result.get('error_description')}")
            
    except Exception as e:
        print(f"Error in email notification: {str(e)}")
    
    print("Email notification process completed")

# Function to validate job data
def validateJobData(job_data):
    available_fields = []
    if 'nodes' in job_data:
        available_fields.append('nodes')
    if 'connections' in job_data:
        available_fields.append('connections')
    if 'metadata' in job_data:
        available_fields.append('metadata')
    
    print(f"Available fields in job data: {', '.join(available_fields)}")
    return True  # Always process the job with whatever data is available

def saveToShapefiles(nodes, connections, anchors, workspace_path):
    """Save nodes, connections, and anchors to shapefiles with WGS 1984 projection."""
    print("\nSaving data to shapefiles...")
    
    # Field name mappings (original -> truncated)
    node_fields = {
        'jobname': 'job_name',
        'job_status': 'job_stat',
        'MR_statu': 'mr_status',
        'company': 'utility',
        'fldcompl': 'completed',
        'tag': 'pole_tag',
        'POA_Height': 'poa_ht',
        'conversation': 'conv',
        'scid': 'scid',
        'last_editor': 'editor',
        'last_edit': 'edit_time'
    }
    
    connection_fields = {
        'connection_id': 'conn_id',
        'connection_type': 'conn_type',
        'node_id_1': 'node1_id',
        'node_id_2': 'node2_id'
    }
    
    anchor_fields = {
        'anchor_spec': 'anch_spec',
        'job_id': 'job_id'
    }
    
    try:
        # Create a timestamp for the master zip file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        master_zip_path = os.path.join(workspace_path, f"KatapultMaster_{timestamp}.zip")
        
        shapefile_components = []  # List to track all shapefile components
        
        # Save and analyze nodes
        if nodes:
            print("\nPole Type Analysis in Shapefile:")
            print("------------------------")
            node_geometries = [Point(node["lng"], node["lat"]) for node in nodes]
            gdf_nodes = gpd.GeoDataFrame(nodes, geometry=node_geometries, crs="EPSG:4326")
            
            # Analyze MR status distribution
            mr_status_counts = gdf_nodes['MR_statu'].value_counts()
            print("\nMR Status Distribution:")
            for status, count in mr_status_counts.items():
                print(f"{status}: {count}")
            
            # Rename columns and drop unnecessary ones
            gdf_nodes.rename(columns=node_fields, inplace=True)
            gdf_nodes.drop(columns=['lat', 'lng', 'pole_class', 'pole_height', 'id'], errors='ignore', inplace=True)
            
            nodes_shp = os.path.join(workspace_path, "poles.shp")
            gdf_nodes.to_file(nodes_shp, driver="ESRI Shapefile")
            print(f"\nPoles shapefile saved successfully with {len(gdf_nodes)} features")
            
            # Track shapefile components
            for ext in ['.shp', '.shx', '.dbf', '.prj', '.cpg']:
                file_path = nodes_shp.replace('.shp', ext)
                if os.path.exists(file_path):
                    shapefile_components.append(file_path)
    
        # Save and analyze connections
        if connections:
            print("\nConnection Type Analysis in Shapefile:")
            print("------------------------")
            valid_connections = []
            line_geometries = []
            connection_types = {}
            
            for connection in connections:
                try:
                    properties = connection.get('properties', {})
                    conn_type = properties.get('connection_type', 'Unknown')
                    if conn_type not in connection_types:
                        connection_types[conn_type] = 0
                    connection_types[conn_type] += 1
                    
                    start_x = properties.get('StartX')
                    start_y = properties.get('StartY')
                    end_x = properties.get('EndX')
                    end_y = properties.get('EndY')
                    
                    if any(coord is None for coord in [start_x, start_y, end_x, end_y]):
                        continue
                        
                    line_geom = LineString([(start_x, start_y), (end_x, end_y)])
                    valid_connections.append(properties)
                    line_geometries.append(line_geom)
                    
                except Exception as e:
                    continue
            
            # Print connection type counts
            for conn_type, count in sorted(connection_types.items()):
                print(f"{conn_type}: {count}")
            
            if valid_connections and line_geometries:
                gdf_connections = gpd.GeoDataFrame(valid_connections, geometry=line_geometries, crs="EPSG:4326")
                gdf_connections.rename(columns=connection_fields, inplace=True)
                
                connections_shp = os.path.join(workspace_path, "connections.shp")
                gdf_connections.to_file(connections_shp, driver="ESRI Shapefile")
                print(f"\nConnections shapefile saved successfully with {len(gdf_connections)} features")
                
                # Track shapefile components
                for ext in ['.shp', '.shx', '.dbf', '.prj', '.cpg']:
                    file_path = connections_shp.replace('.shp', ext)
                    if os.path.exists(file_path):
                        shapefile_components.append(file_path)
    
        # Save and analyze anchors
        if anchors:
            print("\nAnchor Type Analysis in Shapefile:")
            print("------------------------")
            anchor_geometries = [Point(anchor["longitude"], anchor["latitude"]) for anchor in anchors]
            gdf_anchors = gpd.GeoDataFrame(anchors, geometry=anchor_geometries, crs="EPSG:4326")
            
            # Analyze anchor spec distribution
            anchor_spec_counts = gdf_anchors['anchor_spec'].value_counts()
            print("\nAnchor Spec Distribution:")
            for spec, count in anchor_spec_counts.items():
                print(f"{spec}: {count}")
            
            gdf_anchors.rename(columns=anchor_fields, inplace=True)
            gdf_anchors.drop(columns=['latitude', 'longitude'], errors='ignore', inplace=True)
            
            anchors_shp = os.path.join(workspace_path, "anchors.shp")
            gdf_anchors.to_file(anchors_shp, driver="ESRI Shapefile")
            print(f"\nAnchors shapefile saved successfully with {len(gdf_anchors)} features")
            
            # Track shapefile components
            for ext in ['.shp', '.shx', '.dbf', '.prj', '.cpg']:
                file_path = anchors_shp.replace('.shp', ext)
                if os.path.exists(file_path):
                    shapefile_components.append(file_path)
        
        # Create master zip file containing all shapefile components
        with zipfile.ZipFile(master_zip_path, 'w') as master_zip:
            for file_path in shapefile_components:
                if os.path.exists(file_path):
                    master_zip.write(file_path, os.path.basename(file_path))
                    os.remove(file_path)  # Remove the original file after adding to zip
        
        print(f"\nAll shapefiles have been consolidated into: {master_zip_path}")
            
    except Exception as e:
        print(f"Error saving shapefiles: {str(e)}")
        
    print("\nVerification Summary:")
    print("------------------------")
    print(f"Input Counts:")
    print(f"Poles: {len(nodes)}")
    print(f"Connections: {len(connections)}")
    print(f"Anchors: {len(anchors)}")
    print("------------------------")

# Function to get user list from KatapultPro API
def getUserList():
    URL_PATH = '/api/v2/users'
    headers = {}
    user_map = {}

    for attempt in range(5):
        conn = None
        try:
            conn = http.client.HTTPSConnection("katapultpro.com", timeout=10)
            conn.request("GET", f"{URL_PATH}?api_key={CONFIG['API_KEY']}", headers=headers)
            res = conn.getresponse()
            data = res.read().decode("utf-8")
            users_dict = json.loads(data)

            if not isinstance(users_dict, dict):
                raise TypeError(f"Expected a dictionary but received {type(users_dict)}: {users_dict}")

            # Create a mapping of user IDs to full names
            for user_id, user_data in users_dict.items():
                name = user_data.get('name', {})
                full_name = f"{name.get('first', '')} {name.get('last', '')}".strip()
                if not full_name:
                    full_name = user_data.get('email', 'Unknown User')
                user_map[user_id] = full_name

            logging.info(f"Retrieved {len(user_map)} users")
            break

        except (socket.error, OSError) as e:
            print(f"Socket error while getting user list: {e}. Retrying...")
            time.sleep(5)
        except Exception as e:
            print(f"Failed to retrieve user list: {e}")
            break
        finally:
            if conn:
                conn.close()

    return user_map

def test_sharepoint_access():
    """Test SharePoint access using existing credentials."""
    try:
        # Load environment variables
        load_dotenv()
        client_id = os.getenv('AZURE_CLIENT_ID')
        client_secret = os.getenv('AZURE_CLIENT_SECRET')
        tenant_id = os.getenv('AZURE_TENANT_ID')
        user_email = os.getenv('EMAIL_USER')

        # Initialize MSAL client
        authority = f"https://login.microsoftonline.com/{tenant_id}"
        app = msal.ConfidentialClientApplication(
            client_id,
            authority=authority,
            client_credential=client_secret
        )

        # Get access token with SharePoint scope
        scopes = ['https://graph.microsoft.com/.default']
        result = app.acquire_token_silent(scopes, account=None)
        if not result:
            result = app.acquire_token_for_client(scopes)

        if 'access_token' in result:
            # Test SharePoint access using Microsoft Graph API
            headers = {
                'Authorization': f"Bearer {result['access_token']}",
                'Content-Type': 'application/json'
            }
            
            # Make a test request to SharePoint
            graph_endpoint = 'https://graph.microsoft.com/v1.0'
            response = requests.get(
                f"{graph_endpoint}/sites",
                headers=headers
            )
            
            if response.status_code == 200:
                print("Successfully connected to SharePoint")
                return True
            else:
                print(f"Failed to access SharePoint. Status code: {response.status_code}")
                print(f"Response: {response.text}")
                return False
        else:
            print(f"Error getting access token: {result.get('error_description')}")
            return False
            
    except Exception as e:
        print(f"Error testing SharePoint access: {str(e)}")
        return False

def initialize_graph_client():
    """Initialize and return a Microsoft Graph API client"""
    try:
        # Load environment variables
        load_dotenv()
        client_id = os.getenv('AZURE_CLIENT_ID')
        client_secret = os.getenv('AZURE_CLIENT_SECRET')
        tenant_id = os.getenv('AZURE_TENANT_ID')
        
        # Initialize MSAL client
        authority = f"https://login.microsoftonline.com/{tenant_id}"
        app = msal.ConfidentialClientApplication(
            client_id,
            authority=authority,
            client_credential=client_secret
        )
        
        # Get access token
        scopes = ['https://graph.microsoft.com/.default']
        result = app.acquire_token_silent(scopes, account=None)
        if not result:
            result = app.acquire_token_for_client(scopes)
            
        if 'access_token' not in result:
            print(f"Error getting access token: {result.get('error_description')}")
            return None
            
        # Create a requests Session with the token
        session = requests.Session()
        session.headers.update({
            'Authorization': f"Bearer {result['access_token']}",
            'Accept': 'application/json'
        })
        session.base_url = 'https://graph.microsoft.com/v1.0'
        
        # Add method to handle full URLs
        def request_with_base_url(method, url, **kwargs):
            if not url.startswith('http'):
                url = f"{session.base_url}/{url}"
            return session.request(method, url, **kwargs)
            
        session.get = lambda url, **kwargs: request_with_base_url('GET', url, **kwargs)
        session.post = lambda url, **kwargs: request_with_base_url('POST', url, **kwargs)
        session.put = lambda url, **kwargs: request_with_base_url('PUT', url, **kwargs)
        session.patch = lambda url, **kwargs: request_with_base_url('PATCH', url, **kwargs)
        session.delete = lambda url, **kwargs: request_with_base_url('DELETE', url, **kwargs)
        
        return session
        
    except Exception as e:
        print(f"Error initializing Graph client: {str(e)}")
        return None

# Main function to run the job for testing
def main(email_list):
    """Main function to process jobs and generate reports."""
    print("Starting main function...")
    all_jobs = []
    
    # Get user list first
    print("Getting user list...")
    user_map = getUserList()
    print(f"Retrieved {len(user_map)} users")
    
    if TEST_ONLY_SPECIFIC_JOB:
        print(f"Testing specific job with ID: {TEST_JOB_ID}")
        all_jobs = [{'id': TEST_JOB_ID, 'name': 'Test Job'}]
    else:
        print("Getting list of all jobs...")
        all_jobs = getJobList()
        
    all_nodes = []
    all_connections = []
    all_anchors = []
    jobs_summary = []
    
    if not all_jobs:
        print("No jobs found.")
        return
        
    total_jobs = len(all_jobs)
    print(f"Found {total_jobs} jobs to process")
    
    # Test SharePoint access first
    print("\nTesting SharePoint access...")
    sharepoint_access = test_sharepoint_access()
    if not sharepoint_access:
        print("Warning: Could not access SharePoint. Will generate local report only.")
    
    for index, job in enumerate(all_jobs, 1):
        print(f"\n{'='*50}")
        print(f"Processing job {index}/{total_jobs}: {job['name']}")
        print(f"{'='*50}")
        
        job_id = job['id']
        job_name = job['name']
        
        print(f"Fetching data for job: {job_name} (ID: {job_id})")
        job_data = getJobData(job_id)
        
        if job_data and validateJobData(job_data):
            # Extract nodes and connections from job data
            nodes = job_data.get('nodes', {})
            connections = job_data.get('connections', {})
            
            print("Extracting nodes...")
            nodes_data = extractNodes(job_data, job_name, job_id, user_map)  # Pass user_map to extractNodes
            print(f"Found {len(nodes_data)} nodes")
            
            print("Extracting connections...")
            connections_data = extractConnections(connections, job_data.get('nodes', {}))
            
            print("Extracting anchors...")
            anchors = extractAnchors(job_data, job_name, job_id)
            print(f"Found {len(anchors)} anchors")
            
            if nodes_data:
                print("Processing nodes for job summary...")
                all_nodes.extend(nodes_data)
                
                # Calculate field completion percentage
                total_nodes = len(job_data.get('nodes', {}))
                field_completed = sum(1 for node in job_data.get('nodes', {}).values() 
                                   if node.get('attributes', {}).get('field_completed', {}).get('value') == True)
                field_complete_pct = (field_completed / total_nodes * 100) if total_nodes > 0 else 0
                
                # Find most recent editor and edit time
                most_recent_editor = 'Unknown'
                last_edit_time = 'Unknown'
                latest_timestamp = 0
                
                for node in nodes_data:
                    if node.get('last_editor') and node.get('last_edit'):
                        # Parse the timestamp from the format "YYYY-MM-DD HH:MM AM/PM MST"
                        try:
                            edit_time = node['last_edit'].replace(' MST', '')
                            edit_dt = datetime.strptime(edit_time, '%Y-%m-%d %I:%M %p')
                            timestamp = edit_dt.timestamp()
                            
                            if timestamp > latest_timestamp:
                                latest_timestamp = timestamp
                                most_recent_editor = node['last_editor']
                                last_edit_time = node['last_edit']
                        except Exception as e:
                            print(f"Error parsing edit time: {e}")
                            continue
                
                # Calculate trace completion percentage by checking all connections
                total_traces = 0
                completed_traces = 0
                for conn_id, connection in connections.items():
                    sections = connection.get('sections', {}).get('midpoint_section', {})
                    photos = sections.get('photos', {})
                    
                    for photo_id, photo_details in photos.items():
                        if photo_details.get('association') == 'main':
                            main_photo = job_data.get('photos', {}).get(photo_id, {})
                            photofirst_data = main_photo.get('photofirst_data', {})
                            
                            # Check if there are wire or guying traces in the photo
                            if 'wire' in photofirst_data or 'guying' in photofirst_data:
                                total_traces += 1
                                if main_photo.get('tracing_complete', {}).get('auto', False):
                                    completed_traces += 1
                
                trace_complete_pct = (completed_traces / total_traces * 100) if total_traces > 0 else 0
                
                # Get utility from first pole with a company value
                utility = 'Unknown'
                for node_data in job_data.get('nodes', {}).values():
                    company_attr = node_data.get('attributes', {}).get('company', {})
                    # Check all possible sources for company value
                    for source in ['-Imported', 'button_added', 'value', 'auto_calced']:
                        company = company_attr.get(source)
                        if company:
                            utility = company
                            break
                    if utility != 'Unknown':
                        break
                
                # Get last modified date
                last_modified = job_data.get('metadata', {}).get('last_modified', 'Unknown')
                if last_modified != 'Unknown':
                    last_modified = datetime.fromtimestamp(last_modified).strftime('%Y-%m-%d')
                
                # Summarize MR Status counts for the job
                mr_status_counts = {}
                for node in nodes_data:
                    mr_status = node['MR_statu']
                    if mr_status not in mr_status_counts:
                        mr_status_counts[mr_status] = 0
                    mr_status_counts[mr_status] += 1
                
                jobs_summary.append({
                    'job_name': job_name,
                    'job_status': job_data.get('metadata', {}).get('job_status', 'Unknown'),
                    'mr_status_counts': mr_status_counts,
                    'last_modified': last_modified,
                    'field_complete_pct': field_complete_pct,
                    'trace_complete_pct': trace_complete_pct,
                    'utility': utility,
                    'most_recent_editor': most_recent_editor,
                    'last_edit_time': last_edit_time
                })
                print(f"Job summary updated with MR status counts: {mr_status_counts}")
            
            if connections_data:
                all_connections.extend(connections_data)
            if anchors:
                all_anchors.extend(anchors)
        
        print(f"Finished processing job {index}/{total_jobs}")
        if index < total_jobs:
            print("Waiting 2 seconds before next job...")
            time.sleep(2)
    
    print("\nProcessing complete. Saving results...")
    if all_nodes or all_connections or all_anchors:
        print("Saving data to master GeoPackage...")
        workspace_path = CONFIG['WORKSPACE_PATH']
        master_geopkg_filename = "Master.gpkg"
        master_geopkg_path = os.path.join(workspace_path, master_geopkg_filename)
        saveMasterGeoPackage(all_nodes, all_connections, all_anchors, master_geopkg_filename)
        print("Master GeoPackage saved successfully")
        
        # Add call to save shapefiles
        print("\nSaving data to shapefiles...")
        saveToShapefiles(all_nodes, all_connections, all_anchors, workspace_path)
    else:
        print("No data extracted for any job. Nothing to save.")
    
    print("\nGenerating report...")
    report_path = None
    if jobs_summary:
        report_path = create_report(jobs_summary)
        print(f"Report generated at: {report_path}")
        
        print("\nSending email notification...")
        send_email_notification(email_list, report_path)
        print("Email notification process completed")
    else:
        print("No job summary data available. Report not generated.")
    
    print("\nMain function completed")
    return True


if __name__ == "__main__":
    # Email list to notify when the report is done

    email_list = ["brandan.lewis@deeplydigital.com"]
    start_time = time.time()  # Record the start time
    main(email_list)
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Total execution time: {elapsed_time:.2f} seconds")
