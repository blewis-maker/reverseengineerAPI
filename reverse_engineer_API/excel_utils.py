from openpyxl.utils import get_column_letter
from datetime import datetime

def create_summary_sheet(wb, jobs_summary, current_time):
    """Create and populate the Jobs Summary sheet"""
    ws_summary = wb.create_sheet("Jobs Summary")
    
    # Add title and timestamp
    ws_summary['A1'] = "Jobs Summary"
    ws_summary['A2'] = current_time.strftime('%-m/%-d/%Y %-I:%M %p')

    # Define status columns (starting from column C)
    status_columns = [
        'Pending Field Collection',
        'Pending Photo Annotation',
        'Sent to PE',
        'Delivered',
        'Pending EMR',
        'Approved for Construction',
        'As Built',
        'Hold',
        'Electric MR',
        'PCO Required',
        'Pole Count'
    ]

    # Add headers in row 4
    ws_summary['B4'] = "Utility"
    for idx, status in enumerate(status_columns, start=3):
        ws_summary.cell(row=4, column=idx, value=status)

    # Group jobs by utility and calculate totals
    utility_summaries = {}
    for job in jobs_summary:
        utility = job['utility']
        job_status = job.get('job_status', 'Unknown').strip()
        pole_count = sum(job['mr_status_counts'].values())
        electric_mr = job['mr_status_counts'].get('Electric MR', 0)
        pco_required = job['mr_status_counts'].get('PCO Required', 0)

        if utility not in utility_summaries:
            utility_summaries[utility] = {
                'Pending Field Collection': 0,
                'Pending Photo Annotation': 0,
                'Sent to PE': 0,
                'Delivered': 0,
                'Pending EMR': 0,
                'Approved for Construction': 0,
                'As Built': 0,
                'Hold': 0,
                'Electric MR': 0,
                'PCO Required': 0,
                'Pole Count': 0
            }

        # Update status counts
        if job_status in utility_summaries[utility]:
            utility_summaries[utility][job_status] += 1
        utility_summaries[utility]['Electric MR'] += electric_mr
        utility_summaries[utility]['PCO Required'] += pco_required
        utility_summaries[utility]['Pole Count'] += pole_count

    # Sort utilities by pending poles (Pending Field Collection + Pending Photo Annotation)
    sorted_utilities = sorted(
        utility_summaries.items(),
        key=lambda x: (x[1]['Pending Field Collection'] + x[1]['Pending Photo Annotation']),
        reverse=True
    )

    # Write utility summaries to sheet
    row_num = 5  # Start data from row 5
    for utility, summary in sorted_utilities:
        # Write utility name
        ws_summary.cell(row=row_num, column=2, value=utility)

        # Write status counts
        for col_idx, status in enumerate(status_columns, start=3):
            ws_summary.cell(row=row_num, column=col_idx, value=summary[status])

        row_num += 1

    return ws_summary 